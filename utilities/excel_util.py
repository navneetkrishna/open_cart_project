import openpyxl
from openpyxl.styles import PatternFill


def get_sheet_name(file):
    workbook = openpyxl.load_workbook(file)
    # sheet = workbook[sheet_name]
    sheet = workbook.active
    return sheet


def get_row_count(file):
    sheet = get_sheet_name(file)
    return sheet.max_row


def get_column_count(file):
    sheet = get_sheet_name(file)
    return sheet.max_column


def read_data(file, row_num, col_num):
    sheet = get_sheet_name(file)
    return sheet.cell(row_num, col_num).value


def write_data(file, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active
    sheet.cell(row_num, col_num).value = data
    workbook.save(file)


def fill_green_color(file, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    green_fill = PatternFill(start_color='60b212',
                             end_color='60b212',
                             fill_type='solid')
    sheet.cell(row_num, col_num).fill = green_fill
    workbook.save(file)


def fill_red_color(file, row_num, col_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    red_fill = PatternFill(start_color='ff0000',
                           end_color='ff0000',
                           fill_type='solid')
    sheet.cell(row_num, col_num).fill = red_fill
    workbook.save(file)
